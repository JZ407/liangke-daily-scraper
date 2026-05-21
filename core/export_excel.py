"""
一键导出量科网新闻到 Excel（标签展开格式）
用法: python export_excel.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from db import get_session, Article
from datetime import date
import os


def run(output_path=None):
    session = get_session()
    articles = session.query(Article).order_by(Article.id).all()
    if not articles:
        print("No articles found in database.")
        return

    max_tags = max(len(a.tags or []) for a in articles)

    wb = Workbook()
    ws = wb.active
    ws.title = '量科网新闻'

    base_headers = ['ID', '标题', '量科网链接', '参考链接', '原始日期', '量科网日期', '来源域名', '正文', '抓取次数']
    tag_headers = [f'标签{i+1}' for i in range(max_tags)]
    headers = base_headers + tag_headers
    ws.append(headers)

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for a in articles:
        ref_url = a.reference_url if a.reference_url != a.liangke_url else ''
        row = [
            a.id,
            a.title,
            a.liangke_url,
            ref_url,
            a.original_date.strftime('%Y-%m-%d') if a.original_date else '',
            a.liangke_date.strftime('%Y-%m-%d') if a.liangke_date else '',
            a.source_domain or '',
            a.content or '',
            a.fetch_count,
        ]
        tags = a.tags or []
        row.extend(tags)
        row.extend([''] * (max_tags - len(tags)))
        ws.append(row)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 60
    ws.column_dimensions['I'].width = 10
    for i in range(max_tags):
        col = chr(ord('J') + i) if i < 17 else chr(ord('A') + (i - 17) // 26) + chr(ord('A') + (i - 17) % 26)
        ws.column_dimensions[col].width = 15

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    if output_path is None:
        output_path = f'量科网新闻_标签展开_{date.today().strftime("%Y%m%d")}.xlsx'

    wb.save(output_path)
    print(f'Exported {len(articles)} articles to: {output_path}')


if __name__ == '__main__':
    run()
